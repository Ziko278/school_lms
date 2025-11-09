from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime

from .models import Attendance, AttendanceRecord
from .forms import AttendanceForm, AttendanceMarkingForm, AttendanceFilterForm, BulkAttendanceUpdateForm
from courses.models import CourseAllocation, CourseRegistration
from admin_site.models import SystemSettings
from utils.decorators import staff_required, student_required


# ========================== LECTURER ATTENDANCE VIEWS ==========================

@login_required
@staff_required
def attendance_list_view(request):
    """List all attendance sessions"""
    staff = request.user.staff

    # Get all allocations for this lecturer
    allocations = CourseAllocation.objects.filter(lecturer=staff)

    attendances = Attendance.objects.filter(
        course_allocation__in=allocations
    ).select_related('course_allocation__course').order_by('-date')

    # Filter by course
    form = AttendanceFilterForm(request.GET, lecturer=staff)
    if form.is_valid():
        course = form.cleaned_data.get('course')
        if course:
            attendances = attendances.filter(course_allocation=course)

        date_from = form.cleaned_data.get('date_from')
        if date_from:
            attendances = attendances.filter(date__gte=date_from)

        date_to = form.cleaned_data.get('date_to')
        if date_to:
            attendances = attendances.filter(date__lte=date_to)

    # Pagination
    paginator = Paginator(attendances, 20)
    page_number = request.GET.get('page')
    attendances_page = paginator.get_page(page_number)

    context = {
        'title': 'Attendance Sessions',
        'attendances_page': attendances_page,
        'form': form,
    }
    return render(request, 'attendance/attendance_list.html', context)


@login_required
@staff_required
def attendance_create_view(request):
    """Create attendance session and mark attendance"""
    staff = request.user.staff
    settings = SystemSettings.get_instance()

    # Get lecturer's current allocations
    allocations = CourseAllocation.objects.filter(
        lecturer=staff,
        session=settings.current_session,
        semester=settings.current_semester
    ).select_related('course')

    if not allocations.exists():
        messages.warning(request, 'You have no course allocations for the current session/semester.')
        return redirect('attendance:attendance_list')

    if request.method == 'POST':
        form = AttendanceForm(request.POST)

        if form.is_valid():
            attendance = form.save(commit=False)

            # Get course allocation from POST
            allocation_id = request.POST.get('course_allocation')
            try:
                allocation = CourseAllocation.objects.get(
                    id=allocation_id,
                    lecturer=staff
                )
                attendance.course_allocation = allocation
                attendance.save()

                # Get registered students for this course
                students = CourseRegistration.objects.filter(
                    course=allocation.course,
                    session=allocation.session,
                    semester=allocation.semester,
                    status='approved'
                ).select_related('student__user')

                # Process attendance marking
                marked_count = 0
                for registration in students:
                    status = request.POST.get(f'student_{registration.student.id}')
                    if status and status in ['present', 'absent', 'late']:
                        AttendanceRecord.objects.create(
                            attendance=attendance,
                            student=registration.student,
                            status=status
                        )
                        marked_count += 1

                messages.success(request, f'Attendance marked for {marked_count} student(s)!')
                return redirect('attendance:attendance_detail', pk=attendance.id)

            except CourseAllocation.DoesNotExist:
                messages.error(request, 'Invalid course allocation.')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = AttendanceForm()

    context = {
        'title': 'Mark Attendance',
        'form': form,
        'allocations': allocations,
    }
    return render(request, 'attendance/attendance_create.html', context)


@login_required
@staff_required
def attendance_edit_view(request, pk):
    """Edit attendance records"""
    attendance = get_object_or_404(Attendance, pk=pk)

    # Check ownership
    if attendance.course_allocation.lecturer != request.user.staff:
        messages.error(request, 'You do not have permission to edit this attendance.')
        return redirect('attendance:attendance_list')

    # Get attendance records
    records = attendance.records.select_related('student__user').order_by('student__matric_number')

    if request.method == 'POST':
        updated_count = 0

        for record in records:
            new_status = request.POST.get(f'student_{record.student.id}')
            if new_status and new_status in ['present', 'absent', 'late']:
                if record.status != new_status:
                    record.status = new_status
                    record.save()
                    updated_count += 1

        messages.success(request, f'{updated_count} attendance record(s) updated!')
        return redirect('attendance:attendance_detail', pk=pk)

    context = {
        'title': 'Edit Attendance',
        'attendance': attendance,
        'records': records,
    }
    return render(request, 'attendance/attendance_edit.html', context)


@login_required
@staff_required
def attendance_detail_view(request, pk):
    """View attendance details with statistics"""
    attendance = get_object_or_404(Attendance, pk=pk)

    # Check ownership
    if attendance.course_allocation.lecturer != request.user.staff:
        messages.error(request, 'You do not have permission to view this attendance.')
        return redirect('attendance:attendance_list')

    # Get records
    records = attendance.records.select_related('student__user').order_by('student__matric_number')

    # Get statistics
    stats = attendance.get_attendance_stats()

    context = {
        'title': f'Attendance - {attendance.course_allocation.course.code} ({attendance.date})',
        'attendance': attendance,
        'records': records,
        'stats': stats,
    }
    return render(request, 'attendance/attendance_detail.html', context)


@login_required
@staff_required
def attendance_report_view(request):
    """Generate attendance report for a course"""
    staff = request.user.staff
    settings = SystemSettings.get_instance()

    # Get filter parameters
    allocation_id = request.GET.get('allocation')
    export = request.GET.get('export')

    if not allocation_id:
        # Show form to select course
        allocations = CourseAllocation.objects.filter(
            lecturer=staff,
            session=settings.current_session,
            semester=settings.current_semester
        ).select_related('course')

        context = {
            'title': 'Attendance Report',
            'allocations': allocations,
        }
        return render(request, 'attendance/report_select.html', context)

    # Get allocation
    allocation = get_object_or_404(CourseAllocation, id=allocation_id, lecturer=staff)

    # Get all attendance sessions for this course
    attendances = Attendance.objects.filter(
        course_allocation=allocation
    ).order_by('date')

    # Get all registered students
    students = CourseRegistration.objects.filter(
        course=allocation.course,
        session=allocation.session,
        semester=allocation.semester,
        status='approved'
    ).select_related('student__user').order_by('student__matric_number')

    # Build attendance matrix
    attendance_data = []
    for student_reg in students:
        student_data = {
            'student': student_reg.student,
            'records': [],
            'present_count': 0,
            'absent_count': 0,
            'late_count': 0,
            'total_classes': 0,
        }

        for attendance in attendances:
            record = AttendanceRecord.objects.filter(
                attendance=attendance,
                student=student_reg.student
            ).first()

            if record:
                student_data['records'].append(record.status)
                student_data['total_classes'] += 1

                if record.status == 'present':
                    student_data['present_count'] += 1
                elif record.status == 'absent':
                    student_data['absent_count'] += 1
                elif record.status == 'late':
                    student_data['late_count'] += 1
            else:
                student_data['records'].append('N/A')

        # Calculate percentage
        if student_data['total_classes'] > 0:
            student_data['percentage'] = round(
                (student_data['present_count'] / student_data['total_classes']) * 100, 2
            )
        else:
            student_data['percentage'] = 0

        attendance_data.append(student_data)

    # Export to Excel
    if export == 'excel':
        return export_attendance_to_excel(allocation, attendances, attendance_data)

    context = {
        'title': f'Attendance Report - {allocation.course.code}',
        'allocation': allocation,
        'attendances': attendances,
        'attendance_data': attendance_data,
    }
    return render(request, 'attendance/attendance_report.html', context)


def export_attendance_to_excel(allocation, attendances, attendance_data):
    """Export attendance report to Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    # Title
    ws['A1'] = f"Attendance Report - {allocation.course.code}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')

    # Column headers
    headers = ['S/N', 'Matric Number', 'Name', 'Present', 'Absent', 'Late', 'Total Classes', 'Percentage']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill

    # Data rows
    for idx, data in enumerate(attendance_data, start=4):
        ws.cell(row=idx, column=1, value=idx - 3)
        ws.cell(row=idx, column=2, value=data['student'].matric_number)
        ws.cell(row=idx, column=3, value=data['student'].user.get_full_name())
        ws.cell(row=idx, column=4, value=data['present_count'])
        ws.cell(row=idx, column=5, value=data['absent_count'])
        ws.cell(row=idx, column=6, value=data['late_count'])
        ws.cell(row=idx, column=7, value=data['total_classes'])
        ws.cell(row=idx, column=8, value=f"{data['percentage']}%")

    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response[
        'Content-Disposition'] = f'attachment; filename=attendance_{allocation.course.code}_{datetime.now().strftime("%Y%m%d")}.xlsx'

    wb.save(response)
    return response


# ========================== STUDENT ATTENDANCE VIEWS ==========================

@login_required
@student_required
def student_attendance_view(request):
    """View student's own attendance records"""
    student = request.user.student
    settings = SystemSettings.get_instance()

    # Get registered courses
    registrations = CourseRegistration.objects.filter(
        student=student,
        session=settings.current_session,
        semester=settings.current_semester,
        status='approved'
    ).select_related('course')

    # Filter by course
    course_id = request.GET.get('course')

    # Get attendance records
    records = AttendanceRecord.objects.filter(
        student=student
    ).select_related('attendance__course_allocation__course', 'attendance').order_by('-attendance__date')

    if course_id:
        records = records.filter(attendance__course_allocation__course_id=course_id)

    # Calculate statistics per course
    course_stats = {}
    for registration in registrations:
        course = registration.course
        course_records = AttendanceRecord.objects.filter(
            student=student,
            attendance__course_allocation__course=course,
            attendance__course_allocation__session=settings.current_session,
            attendance__course_allocation__semester=settings.current_semester
        )

        total = course_records.count()
        present = course_records.filter(status='present').count()
        absent = course_records.filter(status='absent').count()
        late = course_records.filter(status='late').count()

        percentage = (present / total * 100) if total > 0 else 0

        course_stats[course.id] = {
            'total': total,
            'present': present,
            'absent': absent,
            'late': late,
            'percentage': round(percentage, 2)
        }

    # Pagination
    paginator = Paginator(records, 20)
    page_number = request.GET.get('page')
    records_page = paginator.get_page(page_number)

    context = {
        'title': 'My Attendance',
        'records_page': records_page,
        'registrations': registrations,
        'course_stats': course_stats,
        'selected_course': course_id,
    }
    return render(request, 'attendance/student_attendance.html', context)


# ========================== AJAX VIEWS ==========================

@login_required
@staff_required
@require_http_methods(["POST"])
def mark_attendance_ajax(request):
    """Mark attendance via AJAX"""
    try:
        attendance_id = request.POST.get('attendance_id')
        student_id = request.POST.get('student_id')
        status = request.POST.get('status')

        if status not in ['present', 'absent', 'late']:
            return JsonResponse({
                'success': False,
                'message': 'Invalid status'
            }, status=400)

        attendance = get_object_or_404(Attendance, id=attendance_id)

        # Check ownership
        if attendance.course_allocation.lecturer != request.user.staff:
            return JsonResponse({
                'success': False,
                'message': 'Permission denied'
            }, status=403)

        # Update or create record
        record, created = AttendanceRecord.objects.update_or_create(
            attendance=attendance,
            student_id=student_id,
            defaults={'status': status}
        )

        return JsonResponse({
            'success': True,
            'message': 'Attendance marked successfully'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
@staff_required
@require_http_methods(["POST"])
def update_attendance_status_ajax(request):
    """Update single attendance record status via AJAX"""
    try:
        record_id = request.POST.get('record_id')
        new_status = request.POST.get('status')

        if new_status not in ['present', 'absent', 'late']:
            return JsonResponse({
                'success': False,
                'message': 'Invalid status'
            }, status=400)

        record = get_object_or_404(AttendanceRecord, id=record_id)

        # Check ownership
        if record.attendance.course_allocation.lecturer != request.user.staff:
            return JsonResponse({
                'success': False,
                'message': 'Permission denied'
            }, status=403)

        record.status = new_status
        record.save()

        return JsonResponse({
            'success': True,
            'message': 'Attendance updated successfully'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
@staff_required
@require_http_methods(["GET"])
def get_attendance_stats_ajax(request):
    """Get attendance statistics via AJAX"""
    try:
        attendance_id = request.GET.get('attendance_id')
        attendance = get_object_or_404(Attendance, id=attendance_id)

        # Check ownership
        if attendance.course_allocation.lecturer != request.user.staff:
            return JsonResponse({
                'success': False,
                'message': 'Permission denied'
            }, status=403)

        stats = attendance.get_attendance_stats()

        return JsonResponse({
            'success': True,
            'stats': stats
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)