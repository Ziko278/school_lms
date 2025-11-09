from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.db.models import Count, Sum, Q
from django.utils import timezone
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from .models import SchoolInfo, SystemSettings
from .forms import SchoolInfoForm, SystemSettingsForm
from accounts.models import Student, Staff
from courses.models import CourseRegistration
from payments.models import Payment
from results.models import Result
from academics.models import Session, Semester, Department
from utils.decorators import admin_required


# ========================== ADMIN DASHBOARD ==========================

@login_required
@admin_required
def admin_dashboard_view(request):
    """Admin dashboard with system statistics"""

    # Get current session/semester
    settings = SystemSettings.get_instance()
    current_session = settings.current_session
    current_semester = settings.current_semester

    # Student statistics
    total_students = Student.objects.count()
    pending_students = Student.objects.filter(admission_status='pending').count()
    verified_students = Student.objects.filter(admission_status='verified').count()
    admitted_students = Student.objects.filter(admission_status='admitted').count()

    # Staff statistics
    total_staff = Staff.objects.count()

    # Course registration statistics
    total_registrations = CourseRegistration.objects.filter(
        session=current_session,
        semester=current_semester
    ).count()
    pending_registrations = CourseRegistration.objects.filter(
        session=current_session,
        semester=current_semester,
        status='pending'
    ).count()
    approved_registrations = CourseRegistration.objects.filter(
        session=current_session,
        semester=current_semester,
        status='approved'
    ).count()

    # Payment statistics
    total_revenue = Payment.objects.filter(
        status='success'
    ).aggregate(total=Sum('amount'))['total'] or 0

    pending_payments = Payment.objects.filter(status='pending').count()
    successful_payments = Payment.objects.filter(status='success').count()

    # Recent registrations (last 30 days)
    thirty_days_ago = timezone.now() - timedelta(days=30)
    recent_students = Student.objects.filter(
        created_at__gte=thirty_days_ago
    ).count()

    # Result statistics
    pending_results = Result.objects.filter(status='pending').count()
    verified_results = Result.objects.filter(status='verified').count()

    # Recent activities
    recent_student_registrations = Student.objects.select_related(
        'user', 'department', 'program'
    ).order_by('-created_at')[:5]

    recent_course_registrations = CourseRegistration.objects.select_related(
        'student__user', 'course'
    ).order_by('-registration_date')[:5]

    # Department-wise student count
    department_stats = Department.objects.annotate(
        student_count=Count('students')
    ).order_by('-student_count')[:5]

    context = {
        'title': 'Admin Dashboard',
        'total_students': total_students,
        'pending_students': pending_students,
        'verified_students': verified_students,
        'admitted_students': admitted_students,
        'total_staff': total_staff,
        'total_registrations': total_registrations,
        'pending_registrations': pending_registrations,
        'approved_registrations': approved_registrations,
        'total_revenue': total_revenue,
        'pending_payments': pending_payments,
        'successful_payments': successful_payments,
        'recent_students': recent_students,
        'pending_results': pending_results,
        'verified_results': verified_results,
        'recent_student_registrations': recent_student_registrations,
        'recent_course_registrations': recent_course_registrations,
        'department_stats': department_stats,
        'current_session': current_session,
        'current_semester': current_semester,
    }
    return render(request, 'admin_site/admin_dashboard.html', context)


# ========================== SCHOOL INFO & SETTINGS ==========================

@login_required
@admin_required
def school_info_view(request):
    """View and edit school information"""
    school_info = SchoolInfo.get_instance()

    if request.method == 'POST':
        form = SchoolInfoForm(request.POST, request.FILES, instance=school_info)
        if form.is_valid():
            form.save()
            messages.success(request, 'School information updated successfully!')
            return redirect('admin_site:school_info')
    else:
        form = SchoolInfoForm(instance=school_info)

    context = {
        'title': 'School Information',
        'form': form,
        'school_info': school_info,
    }
    return render(request, 'admin_site/school_info.html', context)


@login_required
@admin_required
def system_settings_view(request):
    """View and edit system settings"""
    settings = SystemSettings.get_instance()

    if request.method == 'POST':
        form = SystemSettingsForm(request.POST, instance=settings)
        if form.is_valid():
            form.save()
            messages.success(request, 'System settings updated successfully!')
            return redirect('admin_site:system_settings')
    else:
        form = SystemSettingsForm(instance=settings)

    context = {
        'title': 'System Settings',
        'form': form,
        'settings': settings,
    }
    return render(request, 'admin_site/system_settings.html', context)


# ========================== REPORTS ==========================

@login_required
@admin_required
def reports_dashboard_view(request):
    """Reports dashboard with links to various reports"""
    context = {
        'title': 'Reports Dashboard',
    }
    return render(request, 'admin_site/reports_dashboard.html', context)


@login_required
@admin_required
def student_enrollment_report_view(request):
    """Generate student enrollment report"""

    # Get filter parameters
    session_id = request.GET.get('session')
    department_id = request.GET.get('department')
    level_id = request.GET.get('level')
    status = request.GET.get('status')
    export = request.GET.get('export')

    # Base queryset
    students = Student.objects.select_related(
        'user', 'department', 'program', 'current_level', 'admission_session'
    ).all()

    # Apply filters
    if session_id:
        students = students.filter(admission_session_id=session_id)
    if department_id:
        students = students.filter(department_id=department_id)
    if level_id:
        students = students.filter(current_level_id=level_id)
    if status:
        students = students.filter(admission_status=status)

    # Statistics
    total_students = students.count()
    male_count = students.filter(user__profile__gender='male').count()
    female_count = students.filter(user__profile__gender='female').count()

    # Department breakdown
    dept_breakdown = students.values(
        'department__name'
    ).annotate(count=Count('id')).order_by('-count')

    # Level breakdown
    level_breakdown = students.values(
        'current_level__name'
    ).annotate(count=Count('id')).order_by('current_level__order')

    # Export to Excel
    if export == 'excel':
        return export_enrollment_to_excel(students, request.GET)

    # Get filter options
    sessions = Session.objects.all()
    departments = Department.objects.all()
    from academics.models import Level
    levels = Level.objects.all()

    context = {
        'title': 'Student Enrollment Report',
        'students': students,
        'total_students': total_students,
        'male_count': male_count,
        'female_count': female_count,
        'dept_breakdown': dept_breakdown,
        'level_breakdown': level_breakdown,
        'sessions': sessions,
        'departments': departments,
        'levels': levels,
        'selected_session': session_id,
        'selected_department': department_id,
        'selected_level': level_id,
        'selected_status': status,
    }
    return render(request, 'admin_site/enrollment_report.html', context)


def export_enrollment_to_excel(students, filters):
    """Export enrollment data to Excel"""

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Enrollment"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Headers
    headers = [
        'S/N', 'Matric Number', 'Full Name', 'Gender', 'Department',
        'Program', 'Level', 'Admission Session', 'Status', 'Email', 'Phone'
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Data rows
    for idx, student in enumerate(students, start=2):
        ws.cell(row=idx, column=1, value=idx - 1)
        ws.cell(row=idx, column=2, value=student.matric_number)
        ws.cell(row=idx, column=3, value=student.user.get_full_name())
        ws.cell(row=idx, column=4, value=student.user.profile.get_gender_display())
        ws.cell(row=idx, column=5, value=student.department.name)
        ws.cell(row=idx, column=6, value=student.program.name)
        ws.cell(row=idx, column=7, value=student.current_level.name)
        ws.cell(row=idx, column=8, value=student.admission_session.name)
        ws.cell(row=idx, column=9, value=student.get_admission_status_display())
        ws.cell(row=idx, column=10, value=student.user.email)
        ws.cell(row=idx, column=11, value=student.user.profile.phone_number)

    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response[
        'Content-Disposition'] = f'attachment; filename=student_enrollment_{datetime.now().strftime("%Y%m%d")}.xlsx'

    wb.save(response)
    return response


@login_required
@admin_required
def payment_report_view(request):
    """Generate payment report"""

    # Get filter parameters
    payment_type = request.GET.get('payment_type')
    status = request.GET.get('status')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    export = request.GET.get('export')

    # Base queryset
    payments = Payment.objects.select_related(
        'student__user', 'admitted_student'
    ).all()

    # Apply filters
    if payment_type:
        payments = payments.filter(payment_type=payment_type)
    if status:
        payments = payments.filter(status=status)
    if date_from:
        payments = payments.filter(created_at__gte=date_from)
    if date_to:
        payments = payments.filter(created_at__lte=date_to)

    # Statistics
    total_payments = payments.count()
    successful_payments = payments.filter(status='success').count()
    pending_payments = payments.filter(status='pending').count()
    failed_payments = payments.filter(status='failed').count()

    total_revenue = payments.filter(status='success').aggregate(
        total=Sum('amount')
    )['total'] or 0

    # Revenue by payment type
    revenue_by_type = payments.filter(status='success').values(
        'payment_type'
    ).annotate(total=Sum('amount')).order_by('-total')

    # Monthly revenue (last 6 months)
    six_months_ago = timezone.now() - timedelta(days=180)
    monthly_revenue = payments.filter(
        status='success',
        payment_date__gte=six_months_ago
    ).extra(
        select={'month': "DATE_TRUNC('month', payment_date)"}
    ).values('month').annotate(
        total=Sum('amount')
    ).order_by('month')

    # Export to Excel
    if export == 'excel':
        return export_payments_to_excel(payments)

    context = {
        'title': 'Payment Report',
        'payments': payments[:100],  # Limit display
        'total_payments': total_payments,
        'successful_payments': successful_payments,
        'pending_payments': pending_payments,
        'failed_payments': failed_payments,
        'total_revenue': total_revenue,
        'revenue_by_type': revenue_by_type,
        'monthly_revenue': monthly_revenue,
        'selected_payment_type': payment_type,
        'selected_status': status,
        'date_from': date_from,
        'date_to': date_to,
        'payment_types': Payment.PAYMENT_TYPE_CHOICES,
        'statuses': Payment.STATUS_CHOICES,
    }
    return render(request, 'admin_site/payment_report.html', context)


def export_payments_to_excel(payments):
    """Export payment data to Excel"""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payments"

    # Headers
    headers = [
        'S/N', 'Reference', 'Student/Name', 'Amount', 'Payment Type',
        'Status', 'Payment Date', 'Payment Method'
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill

    # Data
    for idx, payment in enumerate(payments, start=2):
        student_name = ''
        if payment.student:
            student_name = f"{payment.student.matric_number} - {payment.student.user.get_full_name()}"
        elif payment.admitted_student:
            student_name = f"{payment.admitted_student.first_name} {payment.admitted_student.last_name}"

        ws.cell(row=idx, column=1, value=idx - 1)
        ws.cell(row=idx, column=2, value=payment.reference)
        ws.cell(row=idx, column=3, value=student_name)
        ws.cell(row=idx, column=4, value=float(payment.amount))
        ws.cell(row=idx, column=5, value=payment.get_payment_type_display())
        ws.cell(row=idx, column=6, value=payment.get_status_display())
        ws.cell(row=idx, column=7,
                value=payment.payment_date.strftime('%Y-%m-%d %H:%M') if payment.payment_date else 'N/A')
        ws.cell(row=idx, column=8, value=payment.payment_method)

    # Adjust widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=payment_report_{datetime.now().strftime("%Y%m%d")}.xlsx'

    wb.save(response)
    return response


@login_required
@admin_required
def result_statistics_view(request):
    """Result statistics and analytics"""

    # Get filter parameters
    session_id = request.GET.get('session')
    semester_id = request.GET.get('semester')
    department_id = request.GET.get('department')

    # Base queryset
    results = Result.objects.filter(status='verified')

    # Apply filters
    if session_id:
        results = results.filter(session_id=session_id)
    if semester_id:
        results = results.filter(semester_id=semester_id)
    if department_id:
        results = results.filter(student__department_id=department_id)

    # Overall statistics
    total_results = results.count()

    # Grade distribution
    grade_distribution = results.values('grade').annotate(
        count=Count('id')
    ).order_by('grade')

    # Pass/Fail rate
    pass_count = results.filter(total_score__gte=40).count()
    fail_count = results.filter(total_score__lt=40).count()
    pass_rate = (pass_count / total_results * 100) if total_results > 0 else 0

    # Average score
    from django.db.models import Avg
    avg_score = results.aggregate(avg=Avg('total_score'))['avg'] or 0

    # Department performance
    dept_performance = results.values(
        'student__department__name'
    ).annotate(
        avg_score=Avg('total_score'),
        count=Count('id')
    ).order_by('-avg_score')

    # Top performing students
    from django.db.models import F
    top_students = Student.objects.annotate(
        cgpa=Avg('results__grade_point', filter=Q(results__status='verified'))
    ).filter(cgpa__isnull=False).order_by('-cgpa')[:10]

    # Filter options
    sessions = Session.objects.all()
    semesters = Semester.objects.all()
    departments = Department.objects.all()

    context = {
        'title': 'Result Statistics',
        'total_results': total_results,
        'grade_distribution': grade_distribution,
        'pass_count': pass_count,
        'fail_count': fail_count,
        'pass_rate': round(pass_rate, 2),
        'avg_score': round(avg_score, 2),
        'dept_performance': dept_performance,
        'top_students': top_students,
        'sessions': sessions,
        'semesters': semesters,
        'departments': departments,
        'selected_session': session_id,
        'selected_semester': semester_id,
        'selected_department': department_id,
    }
    return render(request, 'admin_site/result_statistics.html', context)


# ========================== AJAX VIEWS ==========================

@login_required
@admin_required
@require_http_methods(["GET"])
def get_system_stats_ajax(request):
    """Get system statistics via AJAX"""
    try:
        total_students = Student.objects.count()
        total_staff = Staff.objects.count()
        pending_registrations = CourseRegistration.objects.filter(
            status='pending'
        ).count()
        pending_results = Result.objects.filter(status='pending').count()

        settings = SystemSettings.get_instance()

        return JsonResponse({
            'success': True,
            'stats': {
                'total_students': total_students,
                'total_staff': total_staff,
                'pending_registrations': pending_registrations,
                'pending_results': pending_results,
                'current_session': str(settings.current_session) if settings.current_session else 'Not Set',
                'current_semester': str(settings.current_semester) if settings.current_semester else 'Not Set',
            }
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
@admin_required
@require_http_methods(["POST"])
def toggle_registration_ajax(request):
    """Toggle student/course registration"""
    try:
        settings = SystemSettings.get_instance()
        reg_type = request.POST.get('type')

        if reg_type == 'student':
            settings.allow_student_registration = not settings.allow_student_registration
            new_status = settings.allow_student_registration
        elif reg_type == 'course':
            settings.allow_course_registration = not settings.allow_course_registration
            new_status = settings.allow_course_registration
        else:
            return JsonResponse({
                'success': False,
                'message': 'Invalid registration type'
            }, status=400)

        settings.save()

        return JsonResponse({
            'success': True,
            'new_status': new_status,
            'message': f'{"Student" if reg_type == "student" else "Course"} registration {"enabled" if new_status else "disabled"}'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
@admin_required
@require_http_methods(["POST"])
def set_active_session_ajax(request):
    """Set active session"""
    try:
        session_id = request.POST.get('session_id')
        session = get_object_or_404(Session, id=session_id)

        # Activate session
        session.activate()

        # Update system settings
        settings = SystemSettings.get_instance()
        settings.current_session = session
        settings.save()

        return JsonResponse({
            'success': True,
            'session_name': session.name,
            'message': f'Session {session.name} activated successfully'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
@admin_required
@require_http_methods(["POST"])
def set_active_semester_ajax(request):
    """Set active semester"""
    try:
        semester_id = request.POST.get('semester_id')
        semester = get_object_or_404(Semester, id=semester_id)

        # Activate semester
        semester.activate()

        # Update system settings
        settings = SystemSettings.get_instance()
        settings.current_semester = semester
        settings.save()

        return JsonResponse({
            'success': True,
            'semester_name': str(semester),
            'message': f'Semester {semester} activated successfully'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)